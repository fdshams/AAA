# -*- coding: utf-8 -*-
"""
Created on Thu Mar  9 09:29:15 2023
Revised on Sat Jun 15 14:23:35 2024
The code was developed to enter a customer details in AAA.
It uses Selenium for web-scraping to enter the details into fields. The https://demo4.sasradius.com/#/user has an example page for which this code was developed.
The AAA server is an instance of the software SAS4 Radius (demo is available in https://demo4.sasradius.com/#/loginand) and was bought for a firm I worked in.
Disclaimer: I have no affiliation whatsoever with SAS4 Radius.
"""

import re
import ipaddress
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

def configure_browser():
    """Configure and return the Chrome WebDriver."""
    chrome_options = Options()
    chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9332")
    driver_path = "D:/Automation/chromedriver.exe"
    return webdriver.Chrome(driver_path, options=chrome_options)

def get_next_ip_address(ip_assignment_method, workbook_path):
    """Get the next IP address based on the assignment method."""
    if ip_assignment_method == 'auto':
        ip_wb = openpyxl.load_workbook(workbook_path, data_only=True)
        sheet = ip_wb["Private IP address for Customer"]
        last_row = sheet.max_row
        last_ip = sheet.cell(row=last_row, column=1).value
        next_ip = ipaddress.IPv4Address(last_ip) + 1
        return str(next_ip)
    elif ip_assignment_method == 'manual':
        return '172.27.8.120'
    return None

def parse_customer_name(name):
    """Parse the customer's name and return first and last names."""
    name_components = name.split()
    first_name = ''
    last_name = ''
    
    if len(name_components) >= 1:
        first_name = name_components[0].capitalize()
    if len(name_components) >= 2:
        if name_components[0].capitalize() not in ["Ahmad", "Mohammad", "Muhammad"]:
            last_name = name_components[1].capitalize()
    if len(name_components) >= 3:
        first_name = ' '.join(name_components[0:2]).title()
        last_name = name_components[2].title()
    
    return first_name, last_name

def standardize_phone_number(raw_phone):
    """Convert raw phone number to international format."""
    result = re.match(r'^[+91 | 0]*([0-9]{9})$', raw_phone)
    if result:
        return '+91' + result.group(1)
    return None

def generate_password(customer_ID, name):
    """Generate a password using the customer's ID and name."""
    name_parts = re.findall(r'\S+', name)
    comp1, comp2 = '', ''
    
    if name_parts:
        if name_parts[0].capitalize() in ["Ahmad", "Mohammad"]:
            comp1 = name_parts[1].capitalize()
            comp2 = name_parts[-1][0].upper()
        else:
            comp1 = name_parts[0].capitalize()
            comp2 = name_parts[-1][0].upper()
    
    if customer_ID >= 9999:
        d_result = re.findall(r'\d+', str(customer_ID))
        comp4 = d_result[0][0]
        comp6 = d_result[0][-3:]
        return f"{comp1}{comp2}@{comp4}%{comp6}"
    else:
        return f"{comp1}{comp2}@#{customer_ID}"

def register_customer(browser, customer_details):
    """Register a customer in the AAA system."""
    browser.get('AAA_URL')
    
    browser.find_element(By.ID, 'username').send_keys(customer_details['customer_ID'])
    browser.find_element(By.ID, 'password').send_keys(customer_details['password'])
    browser.find_element(By.ID, 'confirm_password').send_keys(customer_details['password'])
    browser.find_element(By.ID, 'firstname').send_keys(customer_details['first_name'])
    
    if customer_details['last_name']:
        browser.find_element(By.ID, 'lastname').send_keys(customer_details['last_name'])
    
    browser.find_element(By.ID, 'phone').send_keys(customer_details['phone'])
    browser.find_element(By.ID, 'city').send_keys(customer_details['city'])
    browser.find_element(By.ID, 'static_ip').send_keys(customer_details['IP'])

def update_ip_pool(ip_assignment_method, workbook_path, next_ip, customer_ID):
    """Update the IP pool with the new IP address if assignment method is 'auto'."""
    if ip_assignment_method == 'auto':
        ip_wb = openpyxl.load_workbook(workbook_path)
        sheet = ip_wb["Private IP address for Customer"]
        last_row = sheet.max_row + 1
        sheet.cell(row=last_row, column=1, value=next_ip)
        sheet.cell(row=last_row, column=2, value='255.255.255.255')
        sheet.cell(row=last_row, column=3, value=customer_ID)
        ip_wb.save(workbook_path)

def main():
    # Customer details
    customer_ID = 11223
    name = 'John Doe John'
    raw_phone = '0123456789'
    city = 'Dubai'
    IP_assignment_method = 'auto'
    workbook_path = r"C:/Users/Me/Documents/IP pool/IP Updated.xlsx"
    
    # Initialize browser
    browser = configure_browser()
    
    # Get IP address
    IP = get_next_ip_address(IP_assignment_method, workbook_path)
    
    # Parse customer name
    first_name, last_name = parse_customer_name(name)
    
    # Standardize phone number
    phone = standardize_phone_number(raw_phone)
    
    # Generate password
    password = generate_password(customer_ID, name)
    
    # Customer details dictionary
    customer_details = {
        'customer_ID': customer_ID,
        'first_name': first_name,
        'last_name': last_name,
        'phone': phone,
        'city': city,
        'IP': IP,
        'password': password
    }
    
    # Register customer in AAA
    register_customer(browser, customer_details)
    
    # Update IP pool
    update_ip_pool(IP_assignment_method, workbook_path, IP, customer_ID)

if __name__ == "__main__":
    main()
